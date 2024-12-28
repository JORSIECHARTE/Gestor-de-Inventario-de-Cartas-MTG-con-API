import os
import requests
import logging
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


logging.basicConfig(filename="errors.log", level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def get_card_data(card_name):
    url = f"https://api.scryfall.com/cards/named?exact={card_name}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        price = data.get("prices", {}).get("usd")
        return {
            "price": float(price) if price and price.replace('.', '', 1).isdigit() else 0.0,
            "rarity": data.get("rarity", "No disponible"),
            "edition": data.get("set_name", "No disponible"),
        }
    else:
        return {"price": 0.0, "rarity": "Error", "edition": "Error"}

excel_path = "MTG_Inventory_Template_Updated.xlsx"

if not os.path.exists(excel_path):
    logging.error(f"El archivo {excel_path} no se encontró. Verifica la ruta o el nombre del archivo.")
    raise FileNotFoundError(f"El archivo {excel_path} no se encontró. Verifica la ruta o el nombre del archivo.")

try:
    wb = load_workbook(excel_path)
except InvalidFileException:
    logging.error(f"El archivo {excel_path} no es un archivo Excel válido o está dañado.")
    raise InvalidFileException(f"El archivo {excel_path} no es un archivo Excel válido o está dañado.")
except PermissionError:
    logging.error(f"No se puede acceder al archivo {excel_path}. Asegúrate de que no esté abierto en otra aplicación.")
    raise PermissionError(f"No se puede acceder al archivo {excel_path}. Asegúrate de que no esté abierto en otra aplicación.")

ws = wb.active

card_names = []
for row in range(2, ws.max_row + 1):
    card_name = ws.cell(row=row, column=1).value
    if card_name is None or card_name.strip() == "":
        logging.warning(f"La celda en la fila {row}, columna 'Nombre de la carta' está vacía. Se omitirá esta fila.")
        print(f"Advertencia: La celda en la fila {row}, columna 'Nombre de la carta' está vacía. Se omitirá esta fila.")
        continue
    card_names.append(card_name)

for row, card_name in enumerate(card_names, start=2):
    try:
        card_data = get_card_data(card_name)
        ws.cell(row=row, column=6, value=card_data["price"])
        ws.cell(row=row, column=3, value=card_data["rarity"])
        ws.cell(row=row, column=2, value=card_data["edition"])
    except Exception as e:
        logging.error(f"Error al procesar la carta '{card_name}' en la fila {row}: {e}")
        print(f"Error al procesar la carta '{card_name}' en la fila {row}: {e}")

output_file = "MTG_Inventory_Template_Updated.xlsx"
try:
    wb.save(output_file)
    print(f"Archivo actualizado y guardado como {output_file}.")
except PermissionError:
    logging.error(f"No se puede guardar el archivo porque está abierto en otra aplicación.")
    print(f"No se puede guardar el archivo porque está abierto en otra aplicación. Cierra el archivo y vuelve a intentarlo.")











